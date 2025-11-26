"""
Content Extraction Service for Various File Types.

Extracts plain text from:
- Google Docs, Sheets, Slides (via export API)
- PDF files (via PyPDF2)
- Word documents (via python-docx)
- Excel spreadsheets (via openpyxl)
- Plain text files

Following SIMPLICITY principle: One method per file type, clear error handling.
"""

import io
from typing import Dict, Any
from pathlib import Path

# Document processing libraries
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook


class ContentExtractor:
    """Extracts text content from various file formats."""

    # MIME type mappings
    GOOGLE_MIME_TYPES = {
        'application/vnd.google-apps.document': 'google_doc',
        'application/vnd.google-apps.spreadsheet': 'google_sheet',
        'application/vnd.google-apps.presentation': 'google_slides'
    }

    FILE_MIME_TYPES = {
        'application/pdf': 'pdf',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document': 'docx',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'xlsx',
        'text/plain': 'text',
        'text/csv': 'csv'
    }

    def __init__(self, drive_service):
        """
        Initialize content extractor with Drive service.

        Args:
            drive_service: DriveService instance for downloading files
        """
        self.drive_service = drive_service

    def extract_content(
        self,
        file_id: str,
        file_metadata: Dict[str, Any]
    ) -> str:
        """
        Extract text content from any supported file type.

        Args:
            file_id: Google Drive file ID
            file_metadata: File metadata dict (must include 'mimeType', 'name')

        Returns:
            Extracted text content as string

        Raises:
            ValueError: If file type is not supported
        """
        mime_type = file_metadata.get('mimeType', '')
        file_name = file_metadata.get('name', 'unknown')

        print(f"📄 Extracting content from: {file_name}")
        print(f"   Type: {mime_type}")

        # Route to appropriate extractor based on MIME type
        if mime_type in self.GOOGLE_MIME_TYPES:
            return self._extract_google_workspace(file_id, mime_type)
        elif mime_type == 'application/pdf':
            return self._extract_pdf(file_id)
        elif mime_type.endswith('.document'):  # Word doc
            return self._extract_docx(file_id)
        elif mime_type.endswith('.sheet'):  # Excel
            return self._extract_xlsx(file_id)
        elif mime_type.startswith('text/'):  # Plain text, CSV, etc.
            return self._extract_text(file_id)
        else:
            raise ValueError(
                f"❌ Unsupported file type: {mime_type}\n"
                f"Supported: Google Docs/Sheets/Slides, PDF, Word, Excel, Text"
            )

    def _extract_google_workspace(self, file_id: str, mime_type: str) -> str:
        """
        Extract text from Google Workspace files (Docs, Sheets, Slides).

        Uses Google Drive export API to convert to plain text.
        """
        file_type = self.GOOGLE_MIME_TYPES[mime_type]

        if file_type == 'google_doc':
            # Export as plain text
            content = self.drive_service.export_google_doc(file_id, 'text/plain')

        elif file_type == 'google_sheet':
            # Export as CSV, then convert to readable format
            csv_content = self.drive_service.export_google_doc(file_id, 'text/csv')
            # Format CSV nicely (add headers, align columns)
            content = self._format_csv(csv_content)

        elif file_type == 'google_slides':
            # Export as plain text (speaker notes + slide text)
            content = self.drive_service.export_google_doc(file_id, 'text/plain')

        else:
            raise ValueError(f"Unknown Google file type: {file_type}")

        print(f"✅ Extracted {len(content)} characters from Google {file_type}")
        return content

    def _extract_pdf(self, file_id: str) -> str:
        """
        Extract text from PDF files using PyPDF2.

        Note: Works with text-based PDFs. Scanned PDFs (images) will return empty text.
        """
        # Download PDF binary content
        pdf_bytes = self.drive_service.download_file(file_id)

        # Extract text from PDF
        pdf_reader = PdfReader(io.BytesIO(pdf_bytes))
        text_parts = []

        for page_num, page in enumerate(pdf_reader.pages, 1):
            page_text = page.extract_text()
            if page_text.strip():
                text_parts.append(f"--- Page {page_num} ---\n{page_text}")

        content = "\n\n".join(text_parts)

        if not content.strip():
            return "⚠️  This PDF appears to be scanned/image-based. No text could be extracted. OCR would be needed."

        print(f"✅ Extracted {len(content)} characters from PDF ({len(pdf_reader.pages)} pages)")
        return content

    def _extract_docx(self, file_id: str) -> str:
        """Extract text from Word documents (.docx) using python-docx."""
        # Download Word document binary content
        docx_bytes = self.drive_service.download_file(file_id)

        # Parse Word document
        doc = Document(io.BytesIO(docx_bytes))

        # Extract all paragraphs
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        content = "\n\n".join(paragraphs)

        print(f"✅ Extracted {len(content)} characters from Word document ({len(paragraphs)} paragraphs)")
        return content

    def _extract_xlsx(self, file_id: str) -> str:
        """
        Extract text from Excel spreadsheets (.xlsx) using openpyxl.

        Converts each sheet to a readable text format.
        """
        # Download Excel spreadsheet binary content
        xlsx_bytes = self.drive_service.download_file(file_id)

        # Parse Excel workbook
        workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)

        sheet_texts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_texts.append(f"=== Sheet: {sheet_name} ===\n")

            # Extract all rows
            for row in sheet.iter_rows(values_only=True):
                # Filter out empty cells and convert to strings
                row_values = [str(cell) for cell in row if cell is not None]
                if row_values:
                    sheet_texts.append(" | ".join(row_values))

            sheet_texts.append("")  # Blank line between sheets

        content = "\n".join(sheet_texts)

        print(f"✅ Extracted {len(content)} characters from Excel ({len(workbook.sheetnames)} sheets)")
        return content

    def _extract_text(self, file_id: str) -> str:
        """Extract content from plain text files (txt, csv, etc.)."""
        # Download as bytes, then decode
        text_bytes = self.drive_service.download_file(file_id)
        content = text_bytes.decode('utf-8')

        print(f"✅ Extracted {len(content)} characters from text file")
        return content

    def _format_csv(self, csv_content: str) -> str:
        """
        Format CSV content for better readability.

        Args:
            csv_content: Raw CSV string

        Returns:
            Formatted CSV with aligned columns
        """
        lines = csv_content.strip().split('\n')

        if not lines:
            return csv_content

        # Simple formatting: just add some visual separation
        formatted_lines = []
        for i, line in enumerate(lines):
            formatted_lines.append(line)
            if i == 0:  # After header row
                formatted_lines.append("-" * 50)

        return "\n".join(formatted_lines)
